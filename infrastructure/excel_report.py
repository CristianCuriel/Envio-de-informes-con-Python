# infrastructure/excel_report_exporter.py

import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class ExcelReportExporter:

    def __init__(self, carpeta_excel='Excel'):
        self.carpeta_excel = carpeta_excel
        os.makedirs(self.carpeta_excel, exist_ok=True)

        self.colores_estado = {
            'Cambio': '93c47d',
            'Reparacion': '93c47d',
            'Nota credito': '93c47d',
            'No aplica a garantia': 'e85c5d',
            'Fuera de garantia': 'e85c5d',
            'devuelto': 'e85c5d',
            'En proceso': 'dfd897',
        }

    def exportar(self, nombre_sede: str, df_sede) -> tuple[bool, str]:

        try: 
            archivo = os.path.join(self.carpeta_excel, f"{nombre_sede}.xlsx")
            df_sede.to_excel(archivo, index=False)
            self._modificar_excel(archivo)
            return True, f"{nombre_sede}.xlsx exportado correctamente."
        except FileExistsError:
            return False, f" El archivo {archivo} ya existe. Por favor, elimínelo o cámbiele el nombre."

        except PermissionError:
            return False, f"No se pudo guardar {archivo} porque está abierto o sin permisos."

        except FileNotFoundError:
            return False, "La ruta de exportación no existe."

        except Exception as e:
            return False, f"Ocurrió un error inesperado: {str(e)}"
        

    def _modificar_excel(self, archivo):
        wb = load_workbook(archivo)
        ws = wb.active

        # Ajustar anchos de columnas
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 2

        encabezados = [cell.value for cell in ws[1]]
        try:
            col_estado_idx = encabezados.index("¿Aplica a cambio, reparacion o NC?") + 1
        except ValueError:
            print(f"No se encontró la columna en {archivo}")
            return

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            cell_estado = row[col_estado_idx - 1]
            cell_a_colorear = row[0]

            if cell_estado.value:
                valor = str(cell_estado.value).strip()
                for clave, color in self.colores_estado.items():
                    if clave.lower() in valor.lower():
                        cell_a_colorear.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        break

        wb.save(archivo)
